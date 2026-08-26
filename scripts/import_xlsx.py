#!/usr/bin/env python3
"""从「Notion 导出的图片 zip + 表格 xlsx」一步入库。

为什么需要这个脚本：
Notion 导出附件时，图片文件名 = 提示词全文，但会把 、 / : 等字符换成 _，
同名附件还会被加上 (1)。所以文件名只能用来「认人」，入库的提示词一律取 xlsx 原文。
匹配方式：两边都去掉标点/空白/下划线后做全文精确比对（只比前几十字会把同头不同尾的升级版弄混）。

备注列的硬规定：
表格里那一列补充 / 备注 / 适用说明是「说明文字」，**不属于提示词正文**，
入库后存在独立的 note 字段，前端也单独一栏展示，永远不能拼进 prompt。
早期版本只找「适用/说明/备注」三个词，而表头写的是「补充」，结果整列静静丢掉了。

用法：
  python3 scripts/import_xlsx.py --zip 附件.zip --xlsx 表格.xlsx            # 追加入库
  python3 scripts/import_xlsx.py --zip a.zip --xlsx a.xlsx --reset        # 先清空旧库
  python3 scripts/import_xlsx.py --zip a.zip --xlsx a.xlsx --dry-run      # 只看匹配结果，不入库
  可选 --note-col 补充（表头名字奇怪时手动指定备注列）
  可选 --titles titles.json（{"1": "标题一", "2": "标题二"}）手工指定标题
"""
import argparse
import json
import re
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
IMG_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff", ".avif"}

# 备注列的别名——写全一点，别再因为表头换个说法就整列丢掉
NOTE_KEYS = ("补充", "备注", "备註", "注释", "注记", "适用", "说明", "说明文字", "remark", "Remark", "note", "Note")


def norm(s: str) -> str:
    """只保留中文/字母/数字，抹平 Notion 对文件名做的字符替换。"""
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]", "", s or "")


def unpack(zip_path: Path, out_dir: Path):
    """按顺序重命名落盘（原文件名太长，系统 unzip 会直接失败）。"""
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True)
    items = []
    with zipfile.ZipFile(zip_path) as z:
        names = [n for n in z.namelist() if not n.endswith("/")]
        names = [n for n in names if Path(n).suffix.lower() in IMG_EXTS and "__MACOSX" not in n]
        for i, name in enumerate(sorted(names), 1):
            dest = out_dir / f"{i:03d}{Path(name).suffix.lower()}"
            with z.open(name) as src, open(dest, "wb") as dst:
                shutil.copyfileobj(src, dst)
            items.append({"n": i, "file": str(dest), "name": Path(name).name})
    return items


def read_xlsx(xlsx_path: Path, note_col: str = None):
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("xlsx 是空的")
    header = [str(c or "") for c in rows[0]]

    def find(*keys, default=None, skip=()):
        for i, h in enumerate(header):
            if i in skip:
                continue
            if any(k in h for k in keys):
                return i
        return default

    i_prompt = find("提示词", "prompt", "Prompt", "文本", default=0)
    i_refs = find("效果", "图片", "参考", skip=(i_prompt,))
    if note_col:
        i_note = find(note_col, skip=(i_prompt, i_refs) if i_refs is not None else (i_prompt,))
        if i_note is None:
            raise SystemExit(f"表头里找不到备注列「{note_col}」；当前表头：{header}")
    else:
        i_note = find(*NOTE_KEYS, skip=(i_prompt, i_refs) if i_refs is not None else (i_prompt,))
    if i_note is None:
        # 兵底：三列表且剩下那列没被认出来，就当备注列，总好过静静丢掉
        rest = [i for i in range(len(header)) if i != i_prompt and i != i_refs]
        if len(rest) == 1:
            i_note = rest[0]

    print(f"· 列识别：提示词=「{header[i_prompt]}」· 参考图=「{header[i_refs] if i_refs is not None else '—'}」· 备注=「{header[i_note] if i_note is not None else '—（本次无备注列）'}」")

    out = []
    for k, r in enumerate(rows[1:], 1):
        prompt = str(r[i_prompt] or "").strip() if i_prompt < len(r) else ""
        if not prompt:
            continue
        refs = str(r[i_refs] or "") if i_refs is not None and i_refs < len(r) else ""
        note = str(r[i_note] or "").strip() if i_note is not None and i_note < len(r) else ""
        out.append({"row": k, "prompt": prompt, "want": len([x for x in refs.split(",") if x.strip()]), "note": note})
    return ws.title, out


def auto_title(row) -> str:
    """从「模板适用」提炼标题；没有就用提示词开头。"""
    src = row["note"] or row["prompt"]
    src = re.sub(r"^\s*GPT[\-\s]?\d*\s*(image\d*)?\s*[:\uff1a]?\s*", "", src, flags=re.I)
    src = re.sub(r"^\s*提示词\s*[:\uff1a]?\s*", "", src)
    src = re.split(r"[\n\(\uff08。]", src.strip())[0].strip()
    src = src.replace(" x ", " × ").replace(" X ", " × ")
    return (src[:34] or "未命名风格").strip(" ·×+")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--zip", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--category", help="不填则用 xlsx 的工作表名")
    ap.add_argument("--default-model", default="通用")
    ap.add_argument("--note-col", help="手动指定备注列的表头关键字（默认自动识别 补充/备注/适用/说明）")
    ap.add_argument("--titles", help="JSON：{\"1\": \"标题\"} 按行号覆盖自动标题")
    ap.add_argument("--reset", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    work = ROOT / "incoming" / Path(args.zip).stem
    imgs = unpack(Path(args.zip), work / "files")
    sheet, rows = read_xlsx(Path(args.xlsx), args.note_col)
    category = (args.category or sheet).strip()
    titles = json.loads(Path(args.titles).read_text(encoding="utf-8")) if args.titles else {}

    index = {}
    for r in rows:
        index.setdefault(norm(r["prompt"]), []).append(r)

    unmatched = []
    for m in imgs:
        stem = re.sub(r"\.[A-Za-z0-9]+$", "", m["name"])
        stem = re.sub(r"\(\d+\)\s*$", "", stem).strip()
        hit = index.get(norm(stem))
        if hit and len(hit) == 1:
            hit[0].setdefault("imgs", []).append(m)
        else:
            unmatched.append(m)

    print(f"工作表「{sheet}」：{len(rows)} 行提示词，附件 {len(imgs)} 张图")
    ok = True
    for r in rows:
        got = len(r.get("imgs", []))
        flag = "  " if (r["want"] in (0, got)) else "!!"
        if flag == "!!":
            ok = False
        mark = " 备" if r["note"] else ""
        print(f"{flag} 行{r['row']:>2}：{got} 张（表格标注 {r['want'] or '?'}）{mark} · {auto_title(r)[:26]}")
    n_note = sum(1 for r in rows if r["note"])
    print(f"· 带补充说明的行：{n_note} / {len(rows)}（补充说明存 note 字段，不进提示词）")
    if unmatched:
        ok = False
        print("✗ 对不上行的图：" + ", ".join(f"{m['n']:03d}" for m in unmatched))
        print("  （常见原因：表格里改过提示词但没重新导出图片，或两个附件混在一起）")
    empty = [r["row"] for r in rows if not r.get("imgs")]
    if empty:
        print("! 没图的行（会跳过）：" + ", ".join(map(str, empty)))

    items = []
    for r in rows:
        shots = sorted(r.get("imgs", []), key=lambda m: m["n"])
        if not shots:
            continue
        note = r["note"]
        model = "gpt2" if re.search(r"GPT", note or "", re.I) else args.default_model
        items.append({
            "title": titles.get(str(r["row"])) or auto_title(r),
            "category": category,
            "kind": category.split("——")[0],
            "model": model,
            "tags": [category.split("——")[0]],
            "prompt": r["prompt"],
            "note": note,
            "images": [m["file"] for m in shots],
        })

    manifest = work / "manifest.json"
    manifest.write_text(json.dumps({"items": items}, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"· 已写清单 {manifest.relative_to(ROOT)}：{len(items)} 条 / {sum(len(i['images']) for i in items)} 张")

    if args.dry_run:
        print("· --dry-run：未入库")
        return 0 if ok else 1

    cmd = [sys.executable, str(ROOT / "scripts" / "ingest.py"), "--manifest", str(manifest)]
    if args.reset:
        cmd.append("--reset")
    rc = subprocess.call(cmd)
    if rc == 0:
        print("· 英文提示词记得补中文对照：python3 scripts/merge_i18n.py --check")
    return rc


if __name__ == "__main__":
    sys.exit(main())
